import os.path
import sys
import re
import fitz
import io
import json
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from urllib.request import urlretrieve
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from io import StringIO
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.discovery import build




class JS:

    def __init__(self, position, name, money, comment, status):
        self.position = position
        self.name = name
        self.money = money
        self.comment = comment
        self.status = status
        self.resume = None

    def search_resume(self):
        resume_path = os.path.realpath(f'{os.getcwd()}/{self.position}/{self.name}')
        # print(resume_path)
        if os.path.exists(f'{resume_path}.doc'):
            self.resume = os.path.abspath(f'{resume_path}.doc')
            # print(self.resume)
        elif os.path.exists(f'{resume_path}.pdf'):
            self.resume = os.path.abspath(f'{resume_path}.pdf')
            # print(self.resume)
        else:
            print('Nope')


class Base:

    def __init__(self, path):
        self.path = path
        self.base = load_workbook(self.path)
        self.all_job_seekers = []

    def all_job_seeker(self):
        all_sheet = self.base.sheetnames
        for i in all_sheet:
            current_sheet = self.base[i]
            row = 2
            # all_job_seeker = []
            while True:
                position = current_sheet.cell(row=row, column=1).value
                if position != None:

                    new_job_seeker = [position,
                                      current_sheet.cell(row=row, column=2).value,
                                      current_sheet.cell(row=row, column=3).value,
                                      current_sheet.cell(row=row, column=4).value,
                                      current_sheet.cell(row=row, column=5).value]
                    self.all_job_seekers.append(new_job_seeker)
                    row += 1
                else:
                    # print(self.all_job_seekers)
                    return False

    def new_job_seeker(self):
        for i in self.all_job_seekers:
            aspirant = JS(*i)
            aspirant.search_resume()
            proc_asp = Processing(aspirant)
            proc_asp.process_name()
            proc_asp.process_contacts()
            proc_asp.get_image()
            proc_asp.get_contacnts()
            proc_asp.get_birthday_date()
            proc_asp.add_in_base()


class Processing(object):
    def __init__(self, object):
        self.comment = object.comment
        self.status = object.status
        self.position = object.position
        self.full_name = object.name
        self.money = object.money
        self.text_resume = ''
        if object.resume != None:
            self.path = os.path.normpath(object.resume)
        # print(self.money)
        self.job_seeker = {}


    def process_name(self):
        full_name = self.full_name.split()
        self.last_name = full_name[0]
        self.first_name = full_name[1]
        # self.job_seeker.update({'last name': self.last_name})
        # self.job_seeker.update({'first name': self.first_name})
        if len(full_name) > 2:
            self.middle_name = full_name[2]
            # self.job_seeker.update({'middle_name': self.middle_name})

    def process_contacts(self):
        if self.path != None:
            if self.path.endswith('.doc'):
                try:
                    SCOPES = ['https://www.googleapis.com/auth/drive']
                    SERVICE_ACCOUNT_FILE = f'{start_dir}\my-python-api-278708-2604d6b4e17e.json'

                    credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
                    service = build('drive', 'v3', credentials=credentials)

                    results = service.files().list(
                        pageSize=100,
                        fields="nextPageToken, files(id, name, mimeType, parents, createdTime)",
                        q="name contains 'data'").execute()
                    file_metadata = {'name': 'test',
                                     'mimeType': 'application/vnd.google-apps.document'
                                     }
                    media = MediaFileUpload(self.path, mimetype='application/msword', resumable=True)
                    file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
                    file_id = file.get('id')
                    request = service.files().export_media(fileId=file_id,
                                                           mimeType='text/html'
                                                           )
                    filename = f'{self.path[:-3]}html'
                    fh = io.FileIO(filename, 'wb')
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while done is False:
                        status, done = downloader.next_chunk()
                        print("Download %d%%." % int(status.progress() * 100))
                    with open(filename, 'r') as f:
                        contents = f.read()
                        soup = BeautifulSoup(contents, 'lxml')
                        text = soup.find_all('p')
                        img = soup.img['src']
                        full_text = ''
                        for i in text:
                            if i.text != '':
                                full_text = full_text + '\n' + i.text
                        self.text_resume = full_text
                        self.image = f'{filename[:-4]}jpg'
                        urlretrieve(img, self.image)
                except:
                    pass

            elif self.path.endswith('.pdf'):
                output_string = StringIO()
                with open(self.path, 'rb') as in_file:
                    parser = PDFParser(in_file)
                    doc = PDFDocument(parser)
                    rsrcmgr = PDFResourceManager()
                    device = TextConverter(rsrcmgr, output_string, laparams=LAParams())
                    interpreter = PDFPageInterpreter(rsrcmgr, device)
                    for page in PDFPage.create_pages(doc):
                        interpreter.process_page(page)
                self.text_resume = output_string.getvalue()
                # self.number = re.search(r'(\+7|8).*?(\d{3}).*?(\d{3}).*?(\d{2}).*?(\d{2})', self.text_resume)
                # regex = re.compile(("([a-z0-9!#$%&'*+\/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+\/=?^_`"
                #                     "{|}~-]+)*(@|\sat\s)(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?(\.|"
                #                     "\sdot\s))+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?)"))
                # self.email = re.search(regex, self.text_resume)

    def get_birthday_date(self):
        full_date = re.search(
            r'(\d{1,2}(\-| |\/)(\d{2}|января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)(\-| |\/)\d{4})',
            self.text_resume)
        # print(full_date)
        if full_date != None:
            date = full_date.group().split()
            self.birthday_day = date[0]
            RU_MONTH_VALUES = {
                'января': 1,
                'февраля': 2,
                'марта': 3,
                'апреля': 4,
                'мая': 5,
                'июня': 6,
                'июля': 7,
                'августа': 8,
                'сентября': 9,
                'октября': 10,
                'ноября': 11,
                'декабря': 12, }
            month = date[1]
            self.birthday_month = RU_MONTH_VALUES.get(month)
            self.birthday_year = date[2]
            # self.job_seeker.update({'birthday_day': self.birthday_day})
            # self.job_seeker.update({'birthday_month': self.birthday_month})
            # self.job_seeker.update({'birthday_year': self.birthday_year})
        else:
            self.birthday_day = None

    def get_contacnts(self):
        self.number = re.search(r'(\+7|8).*?(\d{3}).*?(\d{3}).*?(\d{2}).*?(\d{2})', self.text_resume)
        regex = re.compile(("([a-z0-9!#$%&'*+\/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+\/=?^_`"
                                "{|}~-]+)*(@|\sat\s)(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?(\.|"
                                "\sdot\s))+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?)"))
        self.email = re.search(regex, self.text_resume)
        # if self.email != None:
        #     self.job_seeker.update({'phone': self.number.group(0)})
        #     self.job_seeker.update({'email': self.email.group(0)})

    def get_image(self):
        if self.path != None:
            path = self.path
            if path.endswith('.pdf'):
                pdf = fitz.open(self.path)
                for i in range(len(pdf)):
                    for img in pdf.getPageImageList(i):
                        xref = img[0]
                        pix = fitz.Pixmap(pdf, xref)
                        if pix.n < 5:   # this is GRAY or RGB
                            pix.writePNG(f"{self.path[:-4]}_{i}{xref}.png")
                            self.image = f'{self.path[:-4]}_{i}{xref}.png'
                        else:  # CMYK: convert to RGB first
                            pix1 = fitz.Pixmap(fitz.csRGB, pix)
                            pix1.writePNG(f"{self.path[:-4]}_{i}{xref}.png")
                            self.image = f'{self.path[:-4]}_{i}{xref}.png'
                            pix1 = None
                        pix = None

    def add_in_base(self):
        """
        This code need split on more function, but i haven't time
        """
        global id_vacancy
        self.job_seeker.update({'last_name': self.last_name})
        self.job_seeker.update({'first_name': self.first_name})
        if len(self.full_name.split()) > 2:
            self.job_seeker.update({'middle_name': self.middle_name})
        self.job_seeker.update({"position": self.position})
        self.job_seeker.update({"money": self.money})
        if self.email != None:
            self.job_seeker.update({'phone': self.number.group(0)})
            self.job_seeker.update({'email': self.email.group(0)})
        if self.birthday_day != None:
            self.job_seeker.update({'birthday_day': self.birthday_day})
            self.job_seeker.update({'birthday_month': self.birthday_month})
            self.job_seeker.update({'birthday_year': self.birthday_year})
        self.job_seeker.update({"externals":[{
                    "data": {
                        "body": self.text_resume
                    },
                    "auth_type": "NATIVE",
                    "files": [
                        {
                            "id": 12430
                        }
                    ],
                    "account_source": 119
                }]
        })
        # Added job seeker in base
        head = {
            'User-Agent': 'App/1.0 (test@huntflow.ru)',
            'Host': 'api.huntflow.ru',
            'Authorization': token}
        url = base_url + 'account/6/applicants'
        job_seeker_json = json.dumps(self.job_seeker, ensure_ascii=True)
        response = requests.post(url, headers=head, data=job_seeker_json)
        json_responce = response.json()

        self.man_id = json_responce['id']
        # get vacancy list and id vacancy
        url_vac = base_url + 'account/6/vacancies'
        response = requests.get(url_vac, headers=head)
        json_responce = response.json()
        for i in json_responce.get('items'):
            if self.position == i.get('position'):
                id_vacancy = i.get('id')
                print(id_vacancy)
            # Need added code for create new vacancy
        # get dict with vacancy status
        url_vac_stat = base_url + 'account/6/vacancy/statuses'
        response = requests.get(url_vac_stat, headers=head)
        json_responce = response.json()
        id = []
        name = []
        for i in json_responce.get('items'):
            id.append(i.get('id'))
            name.append(i.get('name'))
        dict_vacancy = dict(zip(name, id))
        # dict for translate vacancy
        dict_status_vacancy = {
            'Отправлено письмо': 'Submitted',
            'Интервью с HR': 'HR Interview',
            'Выставлен оффер': 'Offered',
            'Отказ': 'Declined',
        }
        status = dict_status_vacancy.get(self.status)
        print(status)
        status_id = dict_vacancy.get(status)    # take number status vacancy

        if status == 'Declined':
            add_on_vacancy = {
                'vacancy': id_vacancy,
                'status': status_id,
                'files': [
                    {
                        'id': 1234856
                    }
                ],
                "rejection_reason": self.comment
            }
        else:
            add_on_vacancy = {
                'vacancy': id_vacancy,
                'status': status_id,
                'comment':self.comment,
                'files': [
                    {
                        'id': 1234856
                    }
                ]
            }
        vacancy_add = json.dumps(add_on_vacancy)
        url_add_on_vacancy = f'{base_url}account/6/applicants/{self.man_id}/vacancy'
        response = requests.post(url_add_on_vacancy, headers=head, data=vacancy_add) # add client on vacancy with status
        # json_responce = response.json()
        # if self.image:  # upload image, but not worked. Response [500] {'errors': [{'type': 'server_error'}]}
        #     files = {
        #         'file': (os.path.split(self.image)[1], open(os.path.normpath(self.image), 'rb'))
        #     }
        #     header = {
        #         'User-Agent': 'App/1.0 (test@huntflow.ru)',
        #         'Host': 'api.huntflow.ru',
        #         'Content-Type': 'multipart/form-data',
        #         'X-File-Parce': 'true',
        #         'Authorization': token,
        #     }
        #     url_add_files = base_url + 'account/6/upload'
        #     response = requests.post(url_add_files, headers=header, files=files)
        # if self.path:     # upload resume file(need one function for file and image), but not worked. Response [500] {'errors': [{'type': 'server_error'}]}
        #     files = {
        #         'file': (os.path.split(self.path)[1], open(os.path.normpath(self.path), 'rb'))
        #     }
        #     header = {
        #         'User-Agent': 'App/1.0 (test@huntflow.ru)',
        #         'Host': 'api.huntflow.ru',
        #         'Content-Type': 'multipart/form-data',
        #         'X-File-Parce': 'true',
        #         'Authorization': token,
        #     }
        #     url_add_files = base_url + 'account/6/upload'
        #     response = requests.post(url_add_files, headers=header, files=files)

if __name__ == '__main__':
    if os.path.exists(sys.argv[1]):
        path = sys.argv[1]
        input_token = f'{sys.argv[2]}'
    elif os.path.exists(sys.argv[2]):
        path = sys.argv[2]
        input_token = f'{sys.argv[1]}'
    else:
        sys.exit("Wrong path to base")
    # Need added condition for check token
    start_dir = os.getcwd()
    token = f'Bearer {input_token}'
    base_url = 'https://dev-100-api.huntflow.ru/'
    new_path = os.path.basename(path)
    os.chdir(os.path.dirname(path))
    base = Base(new_path)
    base.all_job_seeker()
    base.new_job_seeker()

